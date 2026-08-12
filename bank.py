import subprocess
import sys
import os
import site
import warnings
import io
import pickle
import re
import tempfile
import hashlib
import time
from datetime import datetime, date, timezone, timedelta
from copy import copy as pycopy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import numbers
import pandas as pd
import numpy as np
import streamlit as st
from dateutil.parser import parse as dt_parse

# ==============================================
# FORCE INSTALL MISSING DEPENDENCIES (failsafe)
# ==============================================
extra_lib = os.path.join(os.getcwd(), 'extra_libs')
os.makedirs(extra_lib, exist_ok=True)
sys.path.insert(0, extra_lib)

required_packages = [
    ('streamlit', 'streamlit'),
    ('pandas', 'pandas'),
    ('numpy', 'numpy'),
    ('openpyxl', 'openpyxl'),
    ('dateutil', 'python-dateutil'),
    ('xlrd', 'xlrd'),
    ('groq', 'groq'),
]

for import_name, pkg_name in required_packages:
    try:
        __import__(import_name)
    except ImportError:
        subprocess.check_call([
            sys.executable, "-m", "pip", "install",
            "--target", extra_lib,
            "--no-cache-dir",
            pkg_name
        ])
        __import__(import_name)

warnings.filterwarnings("ignore")

# =========================
# App config & security
# =========================
APP_VERSION = "3.3.0"
APP_NAME = "Bank & Supplier Reconciliation"
DEPLOYMENT_MODE = os.environ.get("DEPLOYMENT_MODE", "production")
SESSION_TIMEOUT_MINUTES = 60

st.set_page_config(
    page_title=f"{APP_NAME} v{APP_VERSION}",
    page_icon="🏦",
    layout="wide",
    initial_sidebar_state="collapsed",
)

def safe_rerun():
    if hasattr(st, "rerun"):
        st.rerun()
    elif hasattr(st, "experimental_rerun"):
        st.experimental_rerun()

def get_org_password():
    env_pw = os.environ.get("APP_PASSWORD", "").strip()
    if env_pw:
        return env_pw
    try:
        sec_pw = str(st.secrets.get("app_password", "")).strip()
        if sec_pw:
            return sec_pw
    except Exception:
        pass
    return "recon2024"

ORG_PASSWORD = get_org_password()

# =========================
# SPAR Brand Colors
# =========================
SPAR_RED = "#EC1B24"
SPAR_GREEN = "#157946"
SPAR_WHITE = "#FFFFFF"
SPAR_DARK_GREEN = "#0F5C34"

# Theme – combines SPAR red and green
THEME = {
    "bg": "#ffffff",
    "panel": "#ffffff",
    "panel2": "#f7f7f7",
    "text": "#111111",
    "muted": "#5b5b5b",
    "border": "rgba(0,0,0,0.10)",
    "border2": "rgba(0,0,0,0.14)",
    "accent": SPAR_RED,
    "accent2": SPAR_DARK_GREEN,
    "good": SPAR_GREEN,
    "bad": SPAR_RED,
    "neutral": "#6b7280",
}

def apply_style():
    st.markdown(
        f"""
        <style>
        :root {{
            --bg: {THEME['bg']};
            --panel: {THEME['panel']};
            --panel2: {THEME['panel2']};
            --text: {THEME['text']};
            --muted: {THEME['muted']};
            --border: {THEME['border']};
            --border2: {THEME['border2']};
            --accent: {THEME['accent']};
            --accent2: {THEME['accent2']};
            --good: {THEME['good']};
            --bad: {THEME['bad']};
            --neutral: {THEME['neutral']};
        }}
        html {{
            color-scheme: light !important;
        }}
        html, body, [data-testid="stAppViewContainer"], .stApp {{
            background: var(--bg) !important;
            color: var(--text) !important;
        }}
        [data-testid="stHeader"], [data-testid="stToolbar"], #MainMenu, footer {{
            display: none !important;
        }}
        .block-container {{
            max-width: 1400px;
            padding-top: 2.6rem !important;
            padding-bottom: 2.2rem !important;
        }}
        div[data-testid="stFileUploader"] {{
            background: var(--panel) !important;
            border: 1px dashed var(--border2) !important;
            border-radius: 16px !important;
            padding: 10px !important;
            transition: border 0.2s ease;
        }}
        div[data-testid="stFileUploader"]:hover {{
            border: 1px dashed var(--accent) !important;
        }}
        div[data-testid="stFileUploader"] label {{
            color: var(--text) !important;
            font-weight: 800 !important;
            font-size: 14px !important;
        }}
        div[data-testid="stFileUploader"] small {{
            color: var(--muted) !important;
            font-size: 12px !important;
        }}
        .card {{
            background: #ffffff !important;
            border: 1px solid var(--border) !important;
            border-radius: 18px !important;
            padding: 18px 18px !important;
        }}
        .hero {{
            border: 1px solid var(--border) !important;
            border-radius: 22px !important;
            padding: 26px 22px !important;
            background: radial-gradient(900px 260px at 50% -10%, rgba(236,27,36,0.10), transparent 60%),
                        linear-gradient(180deg, #ffffff, #ffffff) !important;
        }}
        .title {{
            font-size: 30px !important;
            font-weight: 800 !important;
            letter-spacing: 0.2px !important;
            margin: 0 !important;
        }}
        .subtitle {{
            margin-top: 8px !important;
            color: var(--muted) !important;
            font-size: 14px !important;
        }}
        .chip {{
            display: inline-flex !important;
            align-items: center !important;
            gap: 8px !important;
            padding: 6px 12px !important;
            border-radius: 999px !important;
            border: 1px solid var(--border) !important;
            background: #ffffff !important;
            font-size: 12px !important;
            font-weight: 650 !important;
            color: var(--muted) !important;
        }}
        .chip-dot {{
            width: 8px !important;
            height: 8px !important;
            border-radius: 999px !important;
            display: inline-block !important;
            background: var(--accent) !important;
        }}
        .metric {{
            border: 1px solid var(--border) !important;
            border-radius: 18px !important;
            padding: 14px 14px !important;
            background: #ffffff !important;
        }}
        .metric-k {{
            font-size: 12px !important;
            color: var(--muted) !important;
            font-weight: 700 !important;
            text-transform: uppercase !important;
        }}
        .metric-v {{
            font-size: 26px !important;
            font-weight: 850 !important;
            margin-top: 6px !important;
            color: var(--good) !important;
        }}
        div.stButton > button {{
            background: var(--accent) !important;
            border: 1px solid var(--accent) !important;
            border-radius: 14px !important;
            padding: 0.7rem 1rem !important;
            font-weight: 750 !important;
            color: #ffffff !important;
        }}
        div.stButton > button:hover {{
            background: var(--accent2) !important;
            border-color: var(--accent2) !important;
        }}
        div[data-baseweb="base-input"] > div,
        div[data-baseweb="input"] > div {{
            background: #ffffff !important;
            border: 1px solid var(--border2) !important;
            border-radius: 14px !important;
        }}
        [data-testid="stDataFrame"] {{
            background: #ffffff !important;
            border: 1px solid var(--border) !important;
            border-radius: 16px !important;
            overflow: hidden !important;
        }}
        a, a:visited {{
            color: var(--accent) !important;
            font-weight: 750 !important;
        }}
        .spar-badge {{
            display: inline-block;
            background: {SPAR_RED};
            color: {SPAR_WHITE};
            font-weight: 900;
            font-size: 24px;
            padding: 12px 24px;
            border-radius: 30px;
            letter-spacing: 3px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.15);
            border: 3px solid {SPAR_GREEN};
        }}
        .spar-login-card {{
            background: {SPAR_WHITE} !important;
            border: 3px solid {SPAR_GREEN} !important;
            border-radius: 24px !important;
            box-shadow: 0 12px 40px rgba(0,0,0,0.10) !important;
        }}
        .spar-login-title {{
            font-size: 32px !important;
            font-weight: 900 !important;
            color: {SPAR_GREEN} !important;
            letter-spacing: 1px !important;
        }}
        .spar-login-sub {{
            font-size: 14px !important;
            color: {SPAR_RED} !important;
            font-weight: 600 !important;
        }}
        .spar-login-btn > button {{
            background: {SPAR_GREEN} !important;
            border: 1px solid {SPAR_GREEN} !important;
            color: {SPAR_WHITE} !important;
            font-weight: 800 !important;
            border-radius: 30px !important;
            padding: 0.8rem 1.5rem !important;
        }}
        .spar-login-btn > button:hover {{
            background: {SPAR_DARK_GREEN} !important;
            border-color: {SPAR_DARK_GREEN} !important;
        }}
        .spar-login-input > div > div {{
            border-radius: 30px !important;
            border: 2px solid {SPAR_GREEN} !important;
        }}
        </style>
        """,
        unsafe_allow_html=True,
    )

apply_style()

# =========================
# Persistent Storage Functions
# =========================
STORAGE_FILE = "recon_data.pkl"

def save_recon_data(data):
    try:
        with open(STORAGE_FILE, "wb") as f:
            pickle.dump(data, f)
        return True
    except Exception:
        return False

def load_recon_data():
    try:
        with open(STORAGE_FILE, "rb") as f:
            return pickle.load(f)
    except FileNotFoundError:
        return None
    except Exception:
        return None

def clear_recon_data():
    try:
        os.remove(STORAGE_FILE)
        return True
    except FileNotFoundError:
        return True
    except Exception:
        return False

# =========================
# Session management (authentication)
# =========================
def touch():
    st.session_state.last_activity = datetime.now()

def is_timed_out():
    last = st.session_state.get("last_activity")
    if not last:
        return False
    return (datetime.now() - last).total_seconds() > SESSION_TIMEOUT_MINUTES * 60

def logout():
    st.session_state.authenticated = False
    for k in list(st.session_state.keys()):
        if k not in ["session_id", "last_activity"]:
            del st.session_state[k]
    safe_rerun()

if "authenticated" not in st.session_state:
    st.session_state.authenticated = False
if "session_id" not in st.session_state:
    st.session_state.session_id = hashlib.sha256(str(time.time()).encode()).hexdigest()[:16]
if "last_activity" not in st.session_state:
    st.session_state.last_activity = datetime.now()
if "reconciliation_done" not in st.session_state:
    st.session_state.reconciliation_done = False
if "bank_df" not in st.session_state:
    st.session_state.bank_df = None
if "ledger_df" not in st.session_state:
    st.session_state.ledger_df = None
if "match_results" not in st.session_state:
    st.session_state.match_results = None
if "working_paper" not in st.session_state:
    st.session_state.working_paper = None
if "recon_statement" not in st.session_state:
    st.session_state.recon_statement = None
if "output_bytes" not in st.session_state:
    st.session_state.output_bytes = None
if "output_filename" not in st.session_state:
    st.session_state.output_filename = None
if "file_info" not in st.session_state:
    st.session_state.file_info = {"bank": None, "ledger": None}
if "opening_balance_manual" not in st.session_state:
    st.session_state.opening_balance_manual = 0.0
if "chat_history" not in st.session_state:
    st.session_state.chat_history = []
if "bank_name" not in st.session_state:
    st.session_state.bank_name = ""

# Load saved data on start
if not st.session_state.reconciliation_done:
    saved = load_recon_data()
    if saved:
        st.session_state.output_bytes = saved.get("output_bytes")
        st.session_state.output_filename = saved.get("output_filename")
        st.session_state.reconciliation_done = True
        st.session_state.match_count = saved.get("match_count", 0)
        st.session_state.working_paper = saved.get("working_paper")
        st.session_state.recon_statement = saved.get("recon_statement")
        st.session_state.match_results = saved.get("match_results")
        if "bank_name" in saved:
            st.session_state.bank_name = saved["bank_name"]

def clear_reconciliation_state():
    st.session_state.reconciliation_done = False
    st.session_state.bank_df = None
    st.session_state.ledger_df = None
    st.session_state.match_results = None
    st.session_state.working_paper = None
    st.session_state.recon_statement = None
    st.session_state.output_bytes = None
    st.session_state.output_filename = None
    st.session_state.file_info = {"bank": None, "ledger": None}
    st.session_state.chat_history = []

# =========================
# SPAR-branded login screen
# =========================
def login_screen():
    st.markdown('<div style="height: 2rem;"></div>', unsafe_allow_html=True)
    c1, c2, c3 = st.columns([1, 1.2, 1])
    with c2:
        st.markdown(
            f"""
            <div class="card spar-login-card" style="margin-top: 2vh; padding: 30px 30px !important; text-align:center;">
                <div style="display:flex; justify-content:center; align-items:center; gap:12px; margin-bottom:10px;">
                    <div class="spar-badge" style="font-size:28px; padding:10px 28px;">SPAR</div>
                </div>
                <div class="spar-login-title">{APP_NAME}</div>
                <div class="spar-login-sub">Sign in to continue.</div>
                <div style="height: 14px;"></div>
                <div style="display:flex; justify-content:center;">
                    <div class="chip" style="border-color: {SPAR_GREEN};">
                        <span class="chip-dot" style="background: {SPAR_GREEN};"></span>
                        Version {APP_VERSION} • {DEPLOYMENT_MODE.title()}
                    </div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        with st.form("login_form", clear_on_submit=True):
            st.markdown('<div class="spar-login-input">', unsafe_allow_html=True)
            pw = st.text_input("Password", type="password", placeholder="Organisation password")
            st.markdown('</div>', unsafe_allow_html=True)
            st.markdown('<div class="spar-login-btn">', unsafe_allow_html=True)
            ok = st.form_submit_button("Sign in", use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)
        if ok:
            if pw == ORG_PASSWORD:
                st.session_state.authenticated = True
                touch()
                saved = load_recon_data()
                if saved:
                    st.session_state.output_bytes = saved.get("output_bytes")
                    st.session_state.output_filename = saved.get("output_filename")
                    st.session_state.reconciliation_done = True
                    st.session_state.match_count = saved.get("match_count", 0)
                    st.session_state.working_paper = saved.get("working_paper")
                    st.session_state.recon_statement = saved.get("recon_statement")
                    st.session_state.match_results = saved.get("match_results")
                    if "bank_name" in saved:
                        st.session_state.bank_name = saved["bank_name"]
                safe_rerun()
            else:
                st.error("Wrong password.")

# =========================
# Authentication check
# =========================
if st.session_state.authenticated and is_timed_out():
    st.session_state.authenticated = False
    st.warning("Session timed out. Sign in again.")
    login_screen()
    st.stop()

if not st.session_state.authenticated:
    login_screen()
    st.stop()

touch()

# =========================
# Top bar with SPAR logo and bank name
# =========================
logo_col, title_col = st.columns([1, 5])
with logo_col:
    st.markdown('<div class="spar-badge">SPAR</div>', unsafe_allow_html=True)
with title_col:
    st.markdown(
        f"""
        <div class="hero" style="text-align:center; border: none; background: transparent; padding: 10px 0;">
            <div class="title">🏦 {APP_NAME} <span style="color: {SPAR_GREEN};">with Chipo</span></div>
            <div class="subtitle">Welcome, Chipo! One app for Bank and Supplier reconciliation.</div>
            <div style="height: 8px;"></div>
            <div style="display:flex; justify-content:center; gap:10px; flex-wrap:wrap;">
                <div class="chip"><span class="chip-dot" style="background: {SPAR_GREEN};"></span> Secure session</div>
                <div class="chip">Session {st.session_state.session_id}</div>
                <div class="chip">Mode {DEPLOYMENT_MODE.title()}</div>
                <div class="chip">Version {APP_VERSION}</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

st.markdown("")

# =========================
# Helper functions (Bank Recon)
# =========================
def to_str(x):
    if pd.isna(x):
        return ""
    return str(x).strip()

def to_num(x):
    if pd.isna(x):
        return np.nan
    if isinstance(x, (int, float, np.number)):
        return float(x)
    s = to_str(x).replace(",", "")
    if '(' in s and ')' in s:
        s = '-' + s.replace('(', '').replace(')', '').strip()
    s = re.sub(r"[^\d\.\-]", "", s)
    if not s or s == '-':
        return np.nan
    try:
        return float(s)
    except Exception:
        return np.nan

def excel_serial_to_date(serial):
    if pd.isna(serial) or not isinstance(serial, (int, float)):
        return pd.NaT
    try:
        base = datetime(1899, 12, 30)
        delta = timedelta(days=float(serial))
        return base + delta
    except:
        return pd.NaT

def to_date(x):
    if isinstance(x, (pd.Timestamp, np.datetime64, datetime, date)):
        return pd.to_datetime(x, errors='coerce')
    if isinstance(x, (int, float, np.number)):
        if x > 10000:
            return excel_serial_to_date(x)
        else:
            try:
                return pd.to_datetime(x, unit='s', errors='coerce')
            except:
                return pd.NaT
    s = to_str(x)
    if not s:
        return pd.NaT
    for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%Y/%m/%d', '%d/%m/%y', '%d-%m-%y']:
        try:
            return pd.to_datetime(s, format=fmt)
        except:
            continue
    try:
        return pd.to_datetime(dt_parse(s, fuzzy=True))
    except Exception:
        return pd.NaT

def format_date(date_val):
    if date_val is None:
        return ''
    if pd.isna(date_val):
        return ''
    if isinstance(date_val, str):
        if date_val == '':
            return ''
        try:
            parsed = pd.to_datetime(date_val, errors='coerce')
            if pd.notna(parsed):
                return parsed.strftime('%d/%m/%y')
            else:
                return date_val[:10] if len(date_val) > 10 else date_val
        except:
            return date_val[:10] if len(date_val) > 10 else date_val
    if isinstance(date_val, (pd.Timestamp, datetime, date)):
        if pd.isna(date_val):
            return ''
        try:
            return date_val.strftime('%d/%m/%y')
        except:
            return ''
    return str(date_val)

def detect_table(df_raw, max_rows=100):
    if df_raw is None or df_raw.empty:
        return None, 0
    best_row = 0
    best_score = 0
    for r in range(min(max_rows, len(df_raw))):
        row = df_raw.iloc[r]
        non_empty = row.notna().sum()
        if non_empty < 2:
            continue
        header_score = 0
        for cell in row:
            if pd.isna(cell):
                continue
            s = str(cell).lower()
            if any(kw in s for kw in ['date', 'amount', 'credit', 'debit', 'ref', 'description', 'posting', 'document', 'narration']):
                header_score += 2
        score = non_empty + header_score
        if score > best_score:
            best_score = score
            best_row = r
    if best_row >= 0 and best_row < len(df_raw):
        headers = df_raw.iloc[best_row].fillna('').astype(str).tolist()
        headers = [h.strip() if h.strip() else f"Column_{i}" for i, h in enumerate(headers)]
        data = df_raw.iloc[best_row + 1:].reset_index(drop=True)
        if data.shape[1] < len(headers):
            for _ in range(len(headers) - data.shape[1]):
                data[f"extra_{_}"] = None
        elif data.shape[1] > len(headers):
            data = data.iloc[:, :len(headers)]
        data.columns = headers
        data = data.dropna(how='all')
        return data, best_row
    return df_raw, 0

def load_bank_statement(file):
    try:
        excel_file = pd.ExcelFile(file)
        sheet_names = excel_file.sheet_names
        visible_sheets = [s for s in sheet_names if not s.startswith('_') and 'hidden' not in s.lower()]
        if not visible_sheets:
            visible_sheets = sheet_names
        df_bank = None
        header_row_used = 0
        for sheet in visible_sheets:
            try:
                df_raw = pd.read_excel(file, sheet_name=sheet, header=None)
                if df_raw is not None and not df_raw.empty:
                    df, header_row = detect_table(df_raw)
                    if df is not None and len(df) > 3:
                        df_bank = df
                        header_row_used = header_row
                        break
            except Exception:
                continue
        if df_bank is None or df_bank.empty:
            df_bank = pd.read_excel(file)
            if df_bank.empty:
                raise ValueError("Could not read bank statement file")
    except Exception:
        df_bank = pd.read_excel(file)

    date_col = None
    credit_col = None
    debit_col = None
    ref_col = None
    desc_col = None
    balance_col = None
    for col in df_bank.columns:
        cl = str(col).lower().strip()
        if 'date' in cl and not date_col:
            date_col = col
        if 'credit' in cl and not credit_col:
            credit_col = col
        if 'debit' in cl and not debit_col:
            debit_col = col
        if 'ref' in cl or 'reference' in cl or 'cheque' in cl:
            ref_col = col
        if 'narrative' in cl or 'description' in cl or 'particulars' in cl or 'transaction' in cl:
            desc_col = col
        if 'balance' in cl and not balance_col:
            balance_col = col
    if not credit_col and not debit_col:
        for col in df_bank.columns:
            if 'amount' in str(col).lower() or 'value' in str(col).lower():
                credit_col = col
                debit_col = col
                break

    opening_balance = None
    for idx, row in df_bank.iterrows():
        row_str = ' '.join(str(v).lower() for v in row.values if pd.notna(v))
        if 'balance at period start' in row_str or 'opening balance' in row_str or 'balance brought forward' in row_str:
            for col in df_bank.columns:
                val = to_num(row[col])
                if not pd.isna(val) and val != 0:
                    opening_balance = val
                    break
            break

    transactions = []
    for idx, row in df_bank.iterrows():
        credit = 0
        debit = 0
        if credit_col and credit_col in df_bank.columns:
            credit = to_num(row[credit_col]) if pd.notna(row[credit_col]) else 0
        if debit_col and debit_col in df_bank.columns:
            debit = to_num(row[debit_col]) if pd.notna(row[debit_col]) else 0
        if credit == 0 and debit == 0 and credit_col == debit_col and credit_col:
            amount = to_num(row[credit_col]) if pd.notna(row[credit_col]) else 0
            if amount == 0:
                continue
            credit = amount if amount > 0 else 0
            debit = -amount if amount < 0 else 0
        if credit == 0 and debit == 0:
            continue
        trans_date = to_date(row[date_col]) if date_col and date_col in df_bank.columns else pd.NaT
        if pd.isna(trans_date):
            continue
        desc = ''
        if desc_col and desc_col in df_bank.columns:
            desc = to_str(row[desc_col])
        if not desc and ref_col and ref_col in df_bank.columns:
            desc = to_str(row[ref_col])
        transactions.append({
            'date': trans_date,
            'reference': to_str(row[ref_col]) if ref_col and ref_col in df_bank.columns else '',
            'description': desc,
            'credit': credit,
            'debit': debit,
            'amount': credit - debit,
            'abs_amount': abs(credit - debit),
            'source': 'BANK'
        })
    df_bank_norm = pd.DataFrame(transactions)
    if df_bank_norm.empty:
        st.error("No valid bank transactions.")
        st.stop()
    closing_balance = None
    if balance_col and balance_col in df_bank.columns:
        last_valid = df_bank[balance_col].dropna().iloc[-1] if not df_bank[balance_col].dropna().empty else None
        closing_balance = to_num(last_valid) if last_valid else None
    return df_bank_norm, opening_balance, closing_balance, header_row_used

def load_ledger(file):
    try:
        excel_file = pd.ExcelFile(file)
        sheet_names = excel_file.sheet_names
        visible_sheets = [s for s in sheet_names if not s.startswith('_') and 'hidden' not in s.lower()]
        if not visible_sheets:
            visible_sheets = sheet_names
        df_ledger = None
        header_row_used = 0
        for sheet in visible_sheets:
            try:
                df_raw = pd.read_excel(file, sheet_name=sheet, header=None)
                if df_raw is not None and not df_raw.empty:
                    df, header_row = detect_table(df_raw)
                    if df is not None and len(df) > 3:
                        df_ledger = df
                        header_row_used = header_row
                        break
            except Exception:
                continue
        if df_ledger is None or df_ledger.empty:
            df_ledger = pd.read_excel(file)
            if df_ledger.empty:
                raise ValueError("Could not read ledger file")
    except Exception:
        df_ledger = pd.read_excel(file)

    date_col = None
    amount_col = None
    desc_col = None
    ref_col = None
    balance_col = None
    for col in df_ledger.columns:
        cl = str(col).lower().strip()
        if 'date' in cl and not date_col:
            date_col = col
        if 'amount' in cl and not amount_col:
            amount_col = col
        if 'description' in cl or 'desc' in cl or 'particulars' in cl or 'details' in cl:
            desc_col = col
        if 'document' in cl or 'ref' in cl or 'external' in cl or 'reference' in cl:
            ref_col = col
        if 'balance' in cl and not balance_col:
            balance_col = col

    transactions = []
    closing_balance = None
    for idx, row in df_ledger.iterrows():
        amount = to_num(row[amount_col]) if amount_col and amount_col in df_ledger.columns else np.nan
        if pd.isna(amount):
            continue
        trans_date = to_date(row[date_col]) if date_col and date_col in df_ledger.columns else pd.NaT
        if pd.isna(trans_date):
            continue
        if balance_col and balance_col in df_ledger.columns:
            running_balance = to_num(row[balance_col])
            if not pd.isna(running_balance):
                closing_balance = running_balance
        transactions.append({
            'date': trans_date,
            'reference': to_str(row[ref_col]) if ref_col and ref_col in df_ledger.columns else '',
            'description': to_str(row[desc_col]) if desc_col and desc_col in df_ledger.columns else '',
            'amount': amount,
            'abs_amount': abs(amount),
            'type': 'CREDIT' if amount > 0 else 'DEBIT',
            'source': 'LEDGER'
        })
    df_ledger_norm = pd.DataFrame(transactions)
    if df_ledger_norm.empty:
        st.error("No valid ledger transactions.")
        st.stop()
    if closing_balance is None:
        closing_balance = df_ledger_norm['amount'].sum() if not df_ledger_norm.empty else 0
    return df_ledger_norm, closing_balance, header_row_used

def match_transactions_bank(bank_df, ledger_df):
    bank_copy = bank_df.copy()
    ledger_copy = ledger_df.copy()
    bank_credits = bank_copy[bank_copy['credit'] > 0].copy() if 'credit' in bank_copy.columns else bank_copy[bank_copy['amount'] > 0].copy()
    bank_debits = bank_copy[bank_copy['debit'] > 0].copy() if 'debit' in bank_copy.columns else bank_copy[bank_copy['amount'] < 0].copy()
    if 'credit' not in bank_copy.columns:
        bank_credits = bank_copy[bank_copy['amount'] > 0].copy()
        bank_debits = bank_copy[bank_copy['amount'] < 0].copy()
    ledger_credits = ledger_copy[ledger_copy['amount'] > 0].copy()
    ledger_debits = ledger_copy[ledger_copy['amount'] < 0].copy()
    bank_credits['abs_amount_rounded'] = bank_credits['abs_amount'].round(2)
    bank_debits['abs_amount_rounded'] = bank_debits['abs_amount'].round(2)
    ledger_credits['abs_amount_rounded'] = ledger_credits['abs_amount'].round(2)
    ledger_debits['abs_amount_rounded'] = ledger_debits['abs_amount'].round(2)
    for df in [bank_credits, bank_debits, ledger_credits, ledger_debits]:
        if not df.empty:
            df['date'] = pd.to_datetime(df['date'])
    bank_credits['id'] = [f'B_C_{i}' for i in range(len(bank_credits))]
    bank_debits['id'] = [f'B_D_{i}' for i in range(len(bank_debits))]
    ledger_credits['id'] = [f'L_C_{i}' for i in range(len(ledger_credits))]
    ledger_debits['id'] = [f'L_D_{i}' for i in range(len(ledger_debits))]
    matches = []
    def match_group(ledger_items, bank_items, txn_type):
        if ledger_items.empty or bank_items.empty:
            return [], [], []
        led_dict = ledger_items.to_dict('records')
        bank_dict = bank_items.to_dict('records')
        matched_ledger_ids = set()
        matched_bank_ids = set()
        group_matches = []
        for l_item in led_dict:
            if l_item['id'] in matched_ledger_ids:
                continue
            for b_item in bank_dict:
                if b_item['id'] in matched_bank_ids:
                    continue
                if abs(l_item['abs_amount_rounded'] - b_item['abs_amount_rounded']) <= 0.01:
                    if l_item['date'].date() == b_item['date'].date():
                        group_matches.append({
                            'ledger_id': l_item['id'],
                            'bank_id': b_item['id'],
                            'amount': l_item['abs_amount_rounded'],
                            'type': txn_type,
                            'match_method': 'amount_and_date',
                            'ledger_date': l_item['date'],
                            'bank_date': b_item['date'],
                            'date_match': True,
                            'ledger_desc': l_item.get('description', ''),
                            'ledger_ref': l_item.get('reference', ''),
                            'bank_desc': b_item.get('description', ''),
                            'bank_ref': b_item.get('reference', '')
                        })
                        matched_ledger_ids.add(l_item['id'])
                        matched_bank_ids.add(b_item['id'])
                        break
        for l_item in led_dict:
            if l_item['id'] in matched_ledger_ids:
                continue
            for b_item in bank_dict:
                if b_item['id'] in matched_bank_ids:
                    continue
                if abs(l_item['abs_amount_rounded'] - b_item['abs_amount_rounded']) <= 0.01:
                    group_matches.append({
                        'ledger_id': l_item['id'],
                        'bank_id': b_item['id'],
                        'amount': l_item['abs_amount_rounded'],
                        'type': txn_type,
                        'match_method': 'amount_only',
                        'ledger_date': l_item['date'],
                        'bank_date': b_item['date'],
                        'date_match': False,
                        'ledger_desc': l_item.get('description', ''),
                        'ledger_ref': l_item.get('reference', ''),
                        'bank_desc': b_item.get('description', ''),
                        'bank_ref': b_item.get('reference', '')
                    })
                    matched_ledger_ids.add(l_item['id'])
                    matched_bank_ids.add(b_item['id'])
                    break
        unmatch_ledger = [item for item in led_dict if item['id'] not in matched_ledger_ids]
        unmatch_bank = [item for item in bank_dict if item['id'] not in matched_bank_ids]
        return group_matches, unmatch_ledger, unmatch_bank
    credit_matches, unmatched_ledger_credits, unmatched_bank_credits = match_group(ledger_credits, bank_credits, 'CREDIT')
    debit_matches, unmatched_ledger_debits, unmatched_bank_debits = match_group(ledger_debits, bank_debits, 'DEBIT')
    matches = credit_matches + debit_matches
    def clean_item(item):
        return {
            'date': item['date'],
            'reference': item.get('reference', ''),
            'description': item.get('description', ''),
            'amount': item['amount'],
            'abs_amount': item.get('abs_amount', abs(item['amount'])),
            'type': item.get('type', 'CREDIT' if item['amount'] > 0 else 'DEBIT'),
            'source': item.get('source', 'LEDGER' if 'id' in item and item['id'].startswith('L') else 'BANK')
        }
    return {
        'matches': matches,
        'unmatched_ledger_credits': [clean_item(x) for x in unmatched_ledger_credits],
        'unmatched_ledger_debits': [clean_item(x) for x in unmatched_ledger_debits],
        'unmatched_bank_credits': [clean_item(x) for x in unmatched_bank_credits],
        'unmatched_bank_debits': [clean_item(x) for x in unmatched_bank_debits]
    }

def build_working_paper_bank(match_results):
    rows = []
    for match in match_results['matches']:
        rows.append({
            'SECTION': 'MATCHED',
            'LEDGER_DATE': match['ledger_date'],
            'LEDGER_DESC': match.get('ledger_desc',''),
            'LEDGER_REF': match.get('ledger_ref',''),
            'LEDGER_AMOUNT': match['amount'] if match['type']=='CREDIT' else -match['amount'],
            'MATCH_STATUS': f"MATCHED - {match['match_method'].replace('_',' ').upper()}",
            'BANK_AMOUNT': match['amount'] if match['type']=='CREDIT' else -match['amount'],
            'BANK_DATE': match['bank_date'],
            'BANK_REF': match.get('bank_ref',''),
            'BANK_DESC': match.get('bank_desc','')
        })
    for item in match_results['unmatched_ledger_credits']:
        rows.append({
            'SECTION': 'UNMATCHED - LEDGER ONLY',
            'LEDGER_DATE': item['date'], 'LEDGER_DESC': item['description'], 'LEDGER_REF': item['reference'],
            'LEDGER_AMOUNT': item['amount'], 'MATCH_STATUS': 'NO BANK MATCH',
            'BANK_AMOUNT': '', 'BANK_DATE': '', 'BANK_REF': '', 'BANK_DESC': ''
        })
    for item in match_results['unmatched_ledger_debits']:
        rows.append({
            'SECTION': 'UNMATCHED - LEDGER ONLY',
            'LEDGER_DATE': item['date'], 'LEDGER_DESC': item['description'], 'LEDGER_REF': item['reference'],
            'LEDGER_AMOUNT': item['amount'], 'MATCH_STATUS': 'NO BANK MATCH',
            'BANK_AMOUNT': '', 'BANK_DATE': '', 'BANK_REF': '', 'BANK_DESC': ''
        })
    for item in match_results['unmatched_bank_credits']:
        rows.append({
            'SECTION': 'UNMATCHED - BANK ONLY',
            'LEDGER_DATE': '', 'LEDGER_DESC': '', 'LEDGER_REF': '', 'LEDGER_AMOUNT': '',
            'MATCH_STATUS': 'NO LEDGER MATCH',
            'BANK_AMOUNT': item['amount'], 'BANK_DATE': item['date'], 'BANK_REF': item['reference'], 'BANK_DESC': item['description']
        })
    for item in match_results['unmatched_bank_debits']:
        rows.append({
            'SECTION': 'UNMATCHED - BANK ONLY',
            'LEDGER_DATE': '', 'LEDGER_DESC': '', 'LEDGER_REF': '', 'LEDGER_AMOUNT': '',
            'MATCH_STATUS': 'NO LEDGER MATCH',
            'BANK_AMOUNT': item['amount'], 'BANK_DATE': item['date'], 'BANK_REF': item['reference'], 'BANK_DESC': item['description']
        })
    return pd.DataFrame(rows)

def build_recon_statement(bank_opening, bank_closing, ledger_closing, match_results, bank_name=""):
    recon_items = []
    total_adjustment = 0.0
    for item in match_results['unmatched_bank_credits']:
        amt = -abs(float(item['amount']))
        recon_items.append({
            'date': item['date'],
            'description': f"BANK ONLY - CREDIT TO POST: {item['description']} ({item['reference']})",
            'adjustment': amt,
            'category': 'Bank-only credit'
        })
        total_adjustment += amt
    for item in match_results['unmatched_bank_debits']:
        amt = abs(float(item['amount']))
        recon_items.append({
            'date': item['date'],
            'description': f"BANK ONLY - DEBIT TO POST: {item['description']} ({item['reference']})",
            'adjustment': amt,
            'category': 'Bank-only debit'
        })
        total_adjustment += amt
    for item in match_results['unmatched_ledger_credits']:
        amt = abs(float(item['amount']))
        recon_items.append({
            'date': item['date'],
            'description': f"LEDGER ONLY - DEPOSIT IN TRANSIT: {item['description']} ({item['reference']})",
            'adjustment': amt,
            'category': 'Ledger-only credit'
        })
        total_adjustment += amt
    for item in match_results['unmatched_ledger_debits']:
        amt = -abs(float(item['amount']))
        recon_items.append({
            'date': item['date'],
            'description': f"LEDGER ONLY - UNPRESENTED PAYMENT: {item['description']} ({item['reference']})",
            'adjustment': amt,
            'category': 'Ledger-only debit'
        })
        total_adjustment += amt
    bank_closing = float(bank_closing or 0)
    ledger_closing = float(ledger_closing or 0)
    adjusted_balance = bank_closing + total_adjustment
    difference = adjusted_balance - ledger_closing
    bank_movement = bank_closing - float(bank_opening or 0)
    ledger_movement = ledger_closing - float(bank_opening or 0)
    return {
        'bank_name': bank_name,
        'opening_balance': float(bank_opening or 0),
        'bank_closing_balance': bank_closing,
        'recon_items': pd.DataFrame(recon_items),
        'total_adjustment': total_adjustment,
        'adjusted_balance': adjusted_balance,
        'ledger_balance': ledger_closing,
        'difference': difference,
        'bank_movement': bank_movement,
        'ledger_movement': ledger_movement,
        'movement_difference': bank_movement - ledger_movement
    }

def export_to_excel_bank(working_paper_df, recon_statement, bank_df, ledger_df, match_results):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        wp_display = working_paper_df.copy()
        wp_display['LEDGER_AMOUNT'] = wp_display['LEDGER_AMOUNT'].apply(
            lambda x: float(x) if pd.notna(x) and x != '' else None
        )
        wp_display['BANK_AMOUNT'] = wp_display['BANK_AMOUNT'].apply(
            lambda x: float(x) if pd.notna(x) and x != '' else None
        )
        wp_display.to_excel(writer, sheet_name='WORKING_PAPER', index=False)
        ws = writer.sheets['WORKING_PAPER']
        date_cols = []
        for col_idx, col_name in enumerate(wp_display.columns, 1):
            if col_name in ['LEDGER_DATE', 'BANK_DATE']:
                date_cols.append(get_column_letter(col_idx))
        for col_letter in date_cols:
            for row in range(2, ws.max_row + 1):
                cell = ws[f"{col_letter}{row}"]
                if cell.value and isinstance(cell.value, datetime):
                    cell.number_format = 'dd/mm/yy'
        amount_cols = []
        for col_idx, col_name in enumerate(wp_display.columns, 1):
            if col_name in ['LEDGER_AMOUNT', 'BANK_AMOUNT']:
                amount_cols.append(get_column_letter(col_idx))
        for col_letter in amount_cols:
            for row in range(2, ws.max_row + 1):
                cell = ws[f"{col_letter}{row}"]
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 50)
        recon_data = [
            ['BANK RECONCILIATION STATEMENT', '', ''],
            ['', '', ''],
            [f'Bank: {recon_statement.get("bank_name", "")}', '', ''],
            ['', '', ''],
            ['Opening bank balance', '', recon_statement['opening_balance']],
            ['Bank closing balance per statement', '', recon_statement['bank_closing_balance']],
            ['', '', ''],
            ['RECONCILING ITEMS', '', ''],
            ['Date', 'Description', 'Adjustment',],
        ]
        for _, item in recon_statement['recon_items'].iterrows():
            recon_data.append([
                item['date'] if pd.notna(item['date']) else '',
                item['description'][:120],
                item['adjustment']
            ])
        recon_data += [
            ['', 'Total reconciling adjustments', recon_statement['total_adjustment']],
            ['', 'Adjusted bank balance', recon_statement['adjusted_balance']],
            ['', 'Ledger closing balance', recon_statement['ledger_balance']],
            ['', 'UNRECONCILED DIFFERENCE', recon_statement['difference']],
            ['', '', ''],
            ['DIAGNOSTIC ANALYSIS', '', ''],
            ['Bank movement during period', recon_statement['bank_movement'], ''],
            ['Ledger movement using same opening basis', recon_statement['ledger_movement'], ''],
            ['Movement difference', recon_statement['movement_difference'], ''],
            ['Opening balance basis', recon_statement['opening_balance'], ''],
            ['', '', ''],
            ['Interpretation', 'Difference must be investigated; do not post a balancing adjustment automatically.', ''],
            ['', '', ''],
            ['Prepared by:', '', ''],
            ['Date:', '', datetime.now().strftime('%d/%m/%y')]
        ]
        recon_df = pd.DataFrame(recon_data)
        recon_df.to_excel(writer, sheet_name='RECON_STATEMENT', index=False, header=False)
        ws_recon = writer.sheets['RECON_STATEMENT']
        for row in ws_recon.iter_rows(min_row=1, max_row=ws_recon.max_row, min_col=3, max_col=3):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
        for row in ws_recon.iter_rows(min_row=9, max_row=ws_recon.max_row, min_col=1, max_col=1):
            for cell in row:
                if isinstance(cell.value, datetime):
                    cell.number_format = 'dd/mm/yy'
        matched_detail = [['Matched Transactions - Detailed View'], ['']]
        matched_detail.append([
            'Ledger Date','Ledger Description','Ledger Ref','Ledger Amount',
            'Bank Date','Bank Ref','Bank Description','Bank Amount','Match Method'
        ])
        for match in match_results['matches']:
            signed_amount = match['amount'] if match['type'] == 'CREDIT' else -match['amount']
            matched_detail.append([
                match['ledger_date'],
                (match.get('ledger_desc','')[:80] if match.get('ledger_desc') else ''),
                (match.get('ledger_ref','')[:40] if match.get('ledger_ref') else ''),
                signed_amount,
                match['bank_date'],
                (match.get('bank_ref','')[:40] if match.get('bank_ref') else ''),
                (match.get('bank_desc','')[:80] if match.get('bank_desc') else ''),
                signed_amount,
                match['match_method']
            ])
        matched_df = pd.DataFrame(matched_detail)
        matched_df.to_excel(writer, sheet_name='MATCHED_DETAIL', index=False, header=False)
        ws_md = writer.sheets['MATCHED_DETAIL']
        for row in ws_md.iter_rows(min_row=3, max_row=ws_md.max_row, min_col=1, max_col=1):
            for cell in row:
                if isinstance(cell.value, datetime):
                    cell.number_format = 'dd/mm/yy'
        for row in ws_md.iter_rows(min_row=3, max_row=ws_md.max_row, min_col=5, max_col=5):
            for cell in row:
                if isinstance(cell.value, datetime):
                    cell.number_format = 'dd/mm/yy'
        for row in ws_md.iter_rows(min_row=3, max_row=ws_md.max_row, min_col=4, max_col=4):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
        for row in ws_md.iter_rows(min_row=3, max_row=ws_md.max_row, min_col=8, max_col=8):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
        uml = [['LEDGER TRANSACTIONS WITH NO BANK MATCH'], [''],
               ['Date','Description','Reference','Amount','Type']]
        for item in match_results['unmatched_ledger_credits']:
            uml.append([item['date'], item['description'][:80], item['reference'][:40],
                        item['amount'], 'CREDIT / DEPOSIT IN TRANSIT'])
        for item in match_results['unmatched_ledger_debits']:
            uml.append([item['date'], item['description'][:80], item['reference'][:40],
                        item['amount'], 'DEBIT / UNPRESENTED PAYMENT'])
        uml_df = pd.DataFrame(uml)
        uml_df.to_excel(writer, sheet_name='UNMATCHED_LEDGER', index=False, header=False)
        ws_ul = writer.sheets['UNMATCHED_LEDGER']
        for row in ws_ul.iter_rows(min_row=3, max_row=ws_ul.max_row, min_col=1, max_col=1):
            for cell in row:
                if isinstance(cell.value, datetime):
                    cell.number_format = 'dd/mm/yy'
        for row in ws_ul.iter_rows(min_row=3, max_row=ws_ul.max_row, min_col=4, max_col=4):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
        umb = [['BANK TRANSACTIONS WITH NO LEDGER MATCH'], [''],
               ['Date','Description','Reference','Amount','Type']]
        for item in match_results['unmatched_bank_credits']:
            umb.append([item['date'], item['description'][:80], item['reference'][:40],
                        item['amount'], 'CREDIT / POST TO LEDGER'])
        for item in match_results['unmatched_bank_debits']:
            umb.append([item['date'], item['description'][:80], item['reference'][:40],
                        item['amount'], 'DEBIT / POST TO LEDGER'])
        umb_df = pd.DataFrame(umb)
        umb_df.to_excel(writer, sheet_name='UNMATCHED_BANK', index=False, header=False)
        ws_ub = writer.sheets['UNMATCHED_BANK']
        for row in ws_ub.iter_rows(min_row=3, max_row=ws_ub.max_row, min_col=1, max_col=1):
            for cell in row:
                if isinstance(cell.value, datetime):
                    cell.number_format = 'dd/mm/yy'
        for row in ws_ub.iter_rows(min_row=3, max_row=ws_ub.max_row, min_col=4, max_col=4):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
        summary_data = [
            ['RECONCILIATION SUMMARY', ''],
            ['', ''],
            ['Bank Name', recon_statement.get('bank_name', '')],
            ['Total bank transactions', len(bank_df)],
            ['Total ledger transactions', len(ledger_df)],
            ['Matched transactions', len(match_results['matches'])],
            ['Unmatched - ledger only',
             len(match_results['unmatched_ledger_credits']) + len(match_results['unmatched_ledger_debits'])],
            ['Unmatched - bank only',
             len(match_results['unmatched_bank_credits']) + len(match_results['unmatched_bank_debits'])],
            ['', ''],
            ['Opening bank balance', recon_statement['opening_balance']],
            ['Bank closing balance', recon_statement['bank_closing_balance']],
            ['Total reconciling adjustments', recon_statement['total_adjustment']],
            ['Adjusted bank balance', recon_statement['adjusted_balance']],
            ['Ledger closing balance', recon_statement['ledger_balance']],
            ['Unreconciled difference', recon_statement['difference']],
            ['', ''],
            ['Status', 'RECONCILED' if abs(recon_statement['difference']) < 0.01 else 'INVESTIGATION REQUIRED']
        ]
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='SUMMARY', index=False, header=False)
        ws_sum = writer.sheets['SUMMARY']
        for row in ws_sum.iter_rows(min_row=4, max_row=ws_sum.max_row, min_col=2, max_col=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
        for sheetname in writer.sheets:
            ws = writer.sheets[sheetname]
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 50)
    return output.getvalue()

# =========================
# AI Assistant (Bank) – keep as is
# =========================
def get_ai_context(match_results, recon_statement):
    context = "Here is the summary of the bank reconciliation:\n\n"
    context += f"Opening bank balance: {recon_statement['opening_balance']:,.2f}\n"
    context += f"Bank closing balance (per statement): {recon_statement['bank_closing_balance']:,.2f}\n"
    context += f"Ledger closing balance: {recon_statement['ledger_balance']:,.2f}\n"
    context += f"Unreconciled difference: {recon_statement['difference']:,.2f}\n"
    context += f"Total adjustments made: {recon_statement['total_adjustment']:,.2f}\n\n"
    context += f"Matched transactions: {len(match_results['matches'])}\n"
    context += f"Unmatched ledger credits: {len(match_results['unmatched_ledger_credits'])}\n"
    context += f"Unmatched ledger debits: {len(match_results['unmatched_ledger_debits'])}\n"
    context += f"Unmatched bank credits: {len(match_results['unmatched_bank_credits'])}\n"
    context += f"Unmatched bank debits: {len(match_results['unmatched_bank_debits'])}\n\n"
    context += "Sample unmatched ledger credits:\n"
    for item in match_results['unmatched_ledger_credits'][:3]:
        context += f"- {item['date'].strftime('%d/%m/%y')}: {item['description']} ({item['reference']}) Amount: {item['amount']:,.2f}\n"
    context += "Sample unmatched ledger debits:\n"
    for item in match_results['unmatched_ledger_debits'][:3]:
        context += f"- {item['date'].strftime('%d/%m/%y')}: {item['description']} ({item['reference']}) Amount: {item['amount']:,.2f}\n"
    context += "Sample unmatched bank credits:\n"
    for item in match_results['unmatched_bank_credits'][:3]:
        context += f"- {item['date'].strftime('%d/%m/%y')}: {item['description']} ({item['reference']}) Amount: {item['amount']:,.2f}\n"
    context += "Sample unmatched bank debits:\n"
    for item in match_results['unmatched_bank_debits'][:3]:
        context += f"- {item['date'].strftime('%d/%m/%y')}: {item['description']} ({item['reference']}) Amount: {item['amount']:,.2f}\n"
    context += "\nMatching logic: first matches by absolute amount, then by exact date when multiple identical amounts exist."
    return context

def get_ai_response(user_question, context, api_key):
    system_prompt = (
        "You are Chipo, a friendly and intelligent assistant specialised in bank reconciliation. "
        "You help users understand their reconciliation results, explain differences, and suggest actions. "
        "You are warm, clear, and professional. Always respond in a helpful and encouraging tone.\n\n"
        "Use the following context to answer the user's questions:\n\n" + context
    )
    if api_key and GROQ_AVAILABLE:
        try:
            client = Groq(api_key=api_key)
            completion = client.chat.completions.create(
                model="llama-3.1-70b-versatile",
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_question}
                ],
                temperature=0.7,
                max_tokens=500,
                stream=False,
            )
            return completion.choices[0].message.content.strip()
        except Exception as e:
            return f"⚠️ API error: {str(e)}\n\nFalling back to rule-based response."
    else:
        return get_rule_based_response(user_question, context)

def get_rule_based_response(user_question, context):
    lower_question = user_question.lower()
    if "difference" in lower_question or ("why" in lower_question and "diff" in lower_question):
        diff_match = re.search(r"Unreconciled difference:\s*([\d,.]+)", context)
        adj_match = re.search(r"Total adjustments made:\s*([\d,.]+)", context)
        diff = diff_match.group(1).replace(',','') if diff_match else "0"
        adj = adj_match.group(1).replace(',','') if adj_match else "0"
        return (f"Chipo says: The difference of {diff} arises because there are unmatched items on both sides.\n"
                "• Unmatched ledger credits (deposits) add to the bank balance.\n"
                "• Unmatched ledger debits (payments) subtract from the bank balance.\n"
                "• Unmatched bank credits add to the ledger, and unmatched bank debits subtract.\n"
                f"Total adjustments made: {adj}.\n\n"
                "Check the unmatched transactions list to identify the specific items. I'm here to help you investigate further!")
    elif "matched" in lower_question and ("count" in lower_question or "how many" in lower_question):
        match = re.search(r"Matched transactions:\s*(\d+)", context)
        if match:
            return f"Chipo says: There are {match.group(1)} matched transactions. That's a good sign – most of your transactions are cleared!"
        else:
            return "Chipo says: I couldn't find the matched count. Please ensure the reconciliation has been run."
    elif "unmatched" in lower_question or "mismatch" in lower_question:
        match_cred = re.search(r"Unmatched ledger credits:\s*(\d+)", context)
        match_deb = re.search(r"Unmatched ledger debits:\s*(\d+)", context)
        match_bcred = re.search(r"Unmatched bank credits:\s*(\d+)", context)
        match_bdeb = re.search(r"Unmatched bank debits:\s*(\d+)", context)
        if match_cred and match_deb and match_bcred and match_bdeb:
            return (f"Chipo says: Unmatched items summary:\n"
                    f"• Ledger credits: {match_cred.group(1)}\n"
                    f"• Ledger debits: {match_deb.group(1)}\n"
                    f"• Bank credits: {match_bcred.group(1)}\n"
                    f"• Bank debits: {match_bdeb.group(1)}\n\n"
                    "These are the transactions that need investigation. I recommend reviewing the largest amounts first.")
        else:
            return "Chipo says: I could not find the unmatched counts. Please make sure the reconciliation data is loaded."
    elif "sample" in lower_question or "example" in lower_question:
        sample_lines = [line for line in context.split('\n') if 'Sample unmatched' in line or ('-' in line and 'Amount:' in line)]
        if sample_lines:
            return "Chipo says: Here are some sample unmatched transactions:\n" + "\n".join(sample_lines[:10])
        else:
            return "Chipo says: No sample data available."
    else:
        return ("Chipo says: I can help you understand the reconciliation results. You can ask about:\n"
                "- Why there is a difference\n"
                "- How many transactions were matched\n"
                "- The number of unmatched transactions by category\n"
                "- Sample unmatched transactions\n\n"
                "If you provide more details, I can give a more specific answer. I'm here to make your reconciliation easy!")

# =========================
# Supplier Reconciliation Engine (Fully AI-like)
# =========================
def detect_header_row_supplier(df_raw, max_scan=80):
    best_score = -1
    best_row = 0
    for r in range(min(max_scan, len(df_raw))):
        row = df_raw.iloc[r]
        non_empty = row.notna().sum()
        if non_empty < 2:
            continue
        header_like = 0
        for v in row:
            if pd.isna(v):
                continue
            s = str(v).strip()
            if len(s) > 1 and re.search(r'[A-Za-z]', s):
                header_like += 1
        row_lower = ' '.join([str(v).lower() for v in row if pd.notna(v)])
        keywords = ['date', 'amount', 'description', 'reference', 'invoice', 'document', 'posting', 'balance', 'due', 'entry']
        keyword_score = sum(3 for kw in keywords if kw in row_lower)
        score = non_empty + header_like * 2 + keyword_score
        if score > best_score:
            best_score = score
            best_row = r
    return best_row

def clean_headers_supplier(headers):
    clean = []
    seen = {}
    for h in headers:
        h = str(h).strip()
        if not h:
            h = "Column"
        h = re.sub(r'[^a-zA-Z0-9_ ]', '_', h)
        h = re.sub(r'\s+', '_', h)
        if h in seen:
            seen[h] += 1
            h = f"{h}_{seen[h]}"
        else:
            seen[h] = 0
        clean.append(h)
    return clean

def extract_table_supplier(df_raw, header_row):
    headers = df_raw.iloc[header_row].fillna('').astype(str).tolist()
    headers = clean_headers_supplier(headers)
    data = df_raw.iloc[header_row + 1:].reset_index(drop=True)
    if data.shape[1] < len(headers):
        for _ in range(len(headers) - data.shape[1]):
            data[f"extra_{_}"] = None
    elif data.shape[1] > len(headers):
        data = data.iloc[:, :len(headers)]
    data.columns = headers
    data = data.dropna(how='all')
    return data, headers

def infer_roles_supplier(df):
    cols = df.columns.tolist()
    roles = {}
    def score_col(role, keywords, sample_size=100):
        best_col = None
        best_score = -1
        for col in cols:
            sample = df[col].dropna().head(sample_size).astype(str)
            if len(sample) == 0:
                continue
            col_lower = col.lower()
            name_score = 2 if any(kw in col_lower for kw in keywords) else 0
            if role == 'date':
                date_like = sum(1 for v in sample if is_date_like_supplier(v)) / len(sample)
                data_score = date_like * 10
            elif role == 'amount':
                num_like = sum(1 for v in sample if is_number_like_supplier(v)) / len(sample)
                data_score = num_like * 10
            elif role == 'reference':
                mixed_ratio = sum(1 for v in sample if re.search(r'[A-Za-z/-]', v)) / len(sample)
                data_score = mixed_ratio * 8
            elif role == 'description':
                avg_len = sample.str.len().mean()
                data_score = min(avg_len / 20, 2) * 5
            else:
                data_score = 0
            score = name_score + data_score
            if score > best_score:
                best_score = score
                best_col = col
        return best_col

    roles['date'] = score_col('date', ['date', 'posting', 'doc date', 'entry date', 'transaction date'])
    roles['amount'] = score_col('amount', ['amount', 'amt', 'value', 'total', 'net', 'lcy', 'balance'])
    roles['reference'] = score_col('reference', ['reference', 'doc', 'document', 'invoice', 'inv', 'order', 'po', 'external', 'number'])
    roles['description'] = score_col('description', ['description', 'desc', 'particulars', 'details', 'narration', 'text'])
    if not roles['amount']:
        for col in cols:
            if df[col].apply(is_number_like_supplier).mean() > 0.5:
                roles['amount'] = col
                break
    if not roles['date']:
        for col in cols:
            if df[col].apply(is_date_like_supplier).mean() > 0.3:
                roles['date'] = col
                break
    return roles

def is_date_like_supplier(s):
    try:
        dt_parse(str(s), fuzzy=True)
        return True
    except:
        return False

def is_number_like_supplier(s):
    try:
        float(re.sub(r'[^\d.\-]', '', str(s)))
        return True
    except:
        return False

def parse_amount_supplier(x):
    try:
        s = str(x).replace(',', '').strip()
        if '(' in s and ')' in s:
            s = '-' + re.sub(r'[()]', '', s)
        s = re.sub(r'[^\d.\-]', '', s)
        if s == '' or s == '-':
            return np.nan
        return float(s)
    except:
        return np.nan

def parse_date_supplier(x):
    try:
        return pd.to_datetime(x, errors='coerce')
    except:
        return pd.NaT

def normalize_transactions_supplier(df, roles):
    date_col = roles.get('date')
    amount_col = roles.get('amount')
    ref_col = roles.get('reference')
    desc_col = roles.get('description')

    if not date_col or not amount_col:
        return pd.DataFrame()

    df_clean = df.copy()
    df_clean['_amount'] = df_clean[amount_col].apply(parse_amount_supplier)
    df_clean['_date'] = df_clean[date_col].apply(parse_date_supplier)
    df_clean = df_clean[df_clean['_date'].notna() & df_clean['_amount'].notna()]

    if df_clean.empty:
        return pd.DataFrame()

    if ref_col:
        df_clean['_ref'] = df_clean[ref_col].astype(str).str.strip()
    else:
        df_clean['_ref'] = ''

    if desc_col:
        df_clean['_desc'] = df_clean[desc_col].astype(str).str.strip()
    else:
        df_clean['_desc'] = ''

    summary_keywords = ['total', 'balance', 'opening', 'closing', 'summary', 'subtotal']
    mask = ~(df_clean['_ref'].str.len() == 0) | ~(df_clean['_desc'].str.lower().str.contains('|'.join(summary_keywords)))
    df_clean = df_clean[mask]

    result = pd.DataFrame({
        'date': df_clean['_date'],
        'amount': df_clean['_amount'],
        'reference': df_clean['_ref'],
        'description': df_clean['_desc'],
    })
    return result

def match_transactions_supplier(supplier_df, ledger_df, tolerance=0.01):
    sup = supplier_df.copy()
    led = ledger_df.copy()
    sup['abs_amount'] = sup['amount'].abs().round(2)
    led['abs_amount'] = led['amount'].abs().round(2)
    sup['date'] = pd.to_datetime(sup['date'])
    led['date'] = pd.to_datetime(led['date'])
    sup['row_id'] = [f'S_{i}' for i in range(len(sup))]
    led['row_id'] = [f'L_{i}' for i in range(len(led))]

    matches = []
    used_sup = set()
    used_led = set()

    all_amounts = set(sup['abs_amount'].unique()).union(set(led['abs_amount'].unique()))
    for amt in sorted(all_amounts):
        sup_group = sup[sup['abs_amount'] == amt]
        led_group = led[led['abs_amount'] == amt]
        if sup_group.empty or led_group.empty:
            continue

        sup_list = sup_group.to_dict('records')
        led_list = led_group.to_dict('records')
        for s in sup_list:
            if s['row_id'] in used_sup:
                continue
            matched = None
            for l in led_list:
                if l['row_id'] in used_led:
                    continue
                if s['date'].date() == l['date'].date():
                    matched = l
                    break
            if matched:
                matches.append({
                    'supplier_id': s['row_id'],
                    'ledger_id': matched['row_id'],
                    'amount': amt,
                    'supplier_date': s['date'],
                    'ledger_date': matched['date'],
                    'supplier_ref': s['reference'],
                    'ledger_ref': matched['reference'],
                    'supplier_desc': s['description'],
                    'ledger_desc': matched['description'],
                    'match_method': 'amount_and_date',
                })
                used_sup.add(s['row_id'])
                used_led.add(matched['row_id'])
            else:
                pass

        sup_remaining = sup_group[~sup_group['row_id'].isin(used_sup)]
        led_remaining = led_group[~led_group['row_id'].isin(used_led)]
        for s in sup_remaining.to_dict('records'):
            if s['row_id'] in used_sup:
                continue
            for l in led_remaining.to_dict('records'):
                if l['row_id'] in used_led:
                    continue
                matches.append({
                    'supplier_id': s['row_id'],
                    'ledger_id': l['row_id'],
                    'amount': amt,
                    'supplier_date': s['date'],
                    'ledger_date': l['date'],
                    'supplier_ref': s['reference'],
                    'ledger_ref': l['reference'],
                    'supplier_desc': s['description'],
                    'ledger_desc': l['description'],
                    'match_method': 'amount_only',
                })
                used_sup.add(s['row_id'])
                used_led.add(l['row_id'])
                break

    unmatched_sup = sup[~sup['row_id'].isin(used_sup)].to_dict('records')
    unmatched_led = led[~led['row_id'].isin(used_led)].to_dict('records')

    return {
        'matches': matches,
        'unmatched_supplier': unmatched_sup,
        'unmatched_ledger': unmatched_led,
    }

def format_date_supplier(dt):
    if pd.isna(dt):
        return ''
    return dt.strftime('%d/%m/%y')

def build_working_paper_supplier(match_results):
    rows = []
    for m in match_results['matches']:
        rows.append({
            'SECTION': 'MATCHED',
            'SUPPLIER_DATE': format_date_supplier(m['supplier_date']),
            'SUPPLIER_REF': m['supplier_ref'],
            'SUPPLIER_DESC': m['supplier_desc'],
            'SUPPLIER_AMOUNT': m['amount'],
            'MATCH_METHOD': m['match_method'],
            'LEDGER_DATE': format_date_supplier(m['ledger_date']),
            'LEDGER_REF': m['ledger_ref'],
            'LEDGER_DESC': m['ledger_desc'],
            'LEDGER_AMOUNT': m['amount'],
        })
    for item in match_results['unmatched_supplier']:
        rows.append({
            'SECTION': 'UNMATCHED - SUPPLIER ONLY',
            'SUPPLIER_DATE': format_date_supplier(item['date']),
            'SUPPLIER_REF': item['reference'],
            'SUPPLIER_DESC': item['description'],
            'SUPPLIER_AMOUNT': item['amount'],
            'MATCH_METHOD': 'NO LEDGER MATCH',
            'LEDGER_DATE': '',
            'LEDGER_REF': '',
            'LEDGER_DESC': '',
            'LEDGER_AMOUNT': '',
        })
    for item in match_results['unmatched_ledger']:
        rows.append({
            'SECTION': 'UNMATCHED - LEDGER ONLY',
            'SUPPLIER_DATE': '',
            'SUPPLIER_REF': '',
            'SUPPLIER_DESC': '',
            'SUPPLIER_AMOUNT': '',
            'MATCH_METHOD': 'NO SUPPLIER MATCH',
            'LEDGER_DATE': format_date_supplier(item['date']),
            'LEDGER_REF': item['reference'],
            'LEDGER_DESC': item['description'],
            'LEDGER_AMOUNT': item['amount'],
        })
    return pd.DataFrame(rows)

def reconcile_supplier(supplier_file, ledger_file, tolerance=0.01):
    try:
        sup_raw = pd.read_excel(supplier_file, header=None)
        led_raw = pd.read_excel(ledger_file, header=None)
    except Exception as e:
        raise ValueError(f"Could not read files: {e}")

    sup_header = detect_header_row_supplier(sup_raw)
    led_header = detect_header_row_supplier(led_raw)

    sup_df, _ = extract_table_supplier(sup_raw, sup_header)
    led_df, _ = extract_table_supplier(led_raw, led_header)

    if sup_df.empty or led_df.empty:
        return None, "No data found after table extraction."

    sup_roles = infer_roles_supplier(sup_df)
    led_roles = infer_roles_supplier(led_df)

    sup_norm = normalize_transactions_supplier(sup_df, sup_roles)
    led_norm = normalize_transactions_supplier(led_df, led_roles)

    if sup_norm.empty or led_norm.empty:
        return None, "No valid transactions found after cleaning."

    match_results = match_transactions_supplier(sup_norm, led_norm, tolerance)

    working_paper = build_working_paper_supplier(match_results)

    summary = {
        'total_supplier': len(sup_norm),
        'total_ledger': len(led_norm),
        'matched': len(match_results['matches']),
        'unmatched_supplier': len(match_results['unmatched_supplier']),
        'unmatched_ledger': len(match_results['unmatched_ledger']),
    }

    return {
        'working_paper': working_paper,
        'match_results': match_results,
        'summary': summary,
        'used_columns': {'supplier': sup_roles, 'ledger': led_roles},
    }, None

def export_to_excel_supplier(working_paper_df, match_results):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        working_paper_df.to_excel(writer, sheet_name='WORKING_PAPER', index=False)
        # Summary
        summary = pd.DataFrame([
            ['Matched', len(match_results['matches'])],
            ['Unmatched Supplier', len(match_results['unmatched_supplier'])],
            ['Unmatched Ledger', len(match_results['unmatched_ledger'])],
        ], columns=['Item', 'Count'])
        summary.to_excel(writer, sheet_name='SUMMARY', index=False)
        # Auto-width
        for sheetname in writer.sheets:
            ws = writer.sheets[sheetname]
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 50)
    return output.getvalue()

# =========================
# MAIN UI – Tabs
# =========================
tab_bank, tab_supplier = st.tabs(["🏦 Bank Reconciliation", "🔄 Supplier Reconciliation"])

with tab_bank:
    st.markdown(
        """
        <div class="hero" style="text-align:center; border: none; background: transparent; padding: 10px 0;">
            <div class="title">🏦 Bank Reconciliation</div>
            <div class="subtitle">Upload Bank Statement and Ledger to reconcile closing balances.</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # Bank Name input
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.subheader("🏦 Bank Name")
    bank_name = st.text_input(
        "Enter the name of the bank to reconcile (e.g., CBZ, CABS, NEDBANK)",
        value=st.session_state.bank_name,
        placeholder="e.g., CBZ Bank – USD"
    )
    st.session_state.bank_name = bank_name.strip()
    st.markdown('</div>', unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        bank_file = st.file_uploader(
            "Bank Statement (Excel file)",
            type=["xlsx", "xls"],
            help="Upload your bank statement Excel file.",
            key="bank_file"
        )
        st.caption("Accepts .xlsx or .xls")
        st.markdown('</div>', unsafe_allow_html=True)

    with col2:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        ledger_file = st.file_uploader(
            "Cashbook / Ledger (Excel file)",
            type=["xlsx", "xls"],
            help="Upload your Yellowcob ledger Excel file.",
            key="ledger_file"
        )
        st.caption("Accepts .xlsx or .xls")
        st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.subheader("⚙️ Reconciliation Settings")
    opening_balance_manual = st.number_input(
        "Bank Opening Balance (manual override, 0 = auto-detect)",
        value=st.session_state.opening_balance_manual,
        step=100.00,
        format="%.2f",
        key="bank_opening_input"
    )
    st.session_state.opening_balance_manual = opening_balance_manual
    st.markdown('</div>', unsafe_allow_html=True)

    col_btn1, col_btn2, col_btn3 = st.columns([1,1,1])
    with col_btn1:
        run_bank_recon = st.button("🔄 RUN BANK RECONCILIATION", use_container_width=True, key="run_bank_recon")
    with col_btn2:
        if st.button("🗑️ Clear Results", use_container_width=True, key="clear_bank"):
            clear_reconciliation_state()
            clear_recon_data()
            st.session_state.file_info = {"bank": None, "ledger": None}
            st.session_state.chat_history = []
            st.session_state.bank_name = ""
            safe_rerun()
    with col_btn3:
        if st.button("📥 Download Last Report", use_container_width=True, disabled=(st.session_state.output_bytes is None), key="download_bank_last"):
            if st.session_state.output_bytes:
                st.download_button(
                    label="Download",
                    data=st.session_state.output_bytes,
                    file_name=st.session_state.output_filename or "bank_reconciliation.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="download_bank_last_btn"
                )

    if run_bank_recon:
        if not st.session_state.bank_name:
            st.error("⚠️ Please enter the Bank Name before running reconciliation.")
            st.stop()
        if not bank_file or not ledger_file:
            st.error("Please upload both Bank Statement and Ledger files")
            st.stop()
        with st.spinner("Processing..."):
            try:
                bank_df, bank_opening_auto, bank_closing_auto, _ = load_bank_statement(bank_file)
                ledger_df, ledger_closing, _ = load_ledger(ledger_file)
                opening_balance = st.session_state.opening_balance_manual if st.session_state.opening_balance_manual != 0 else (bank_opening_auto or 0)
                match_results = match_transactions_bank(bank_df, ledger_df)
                working_paper = build_working_paper_bank(match_results)
                recon_statement = build_recon_statement(
                    opening_balance,
                    bank_closing_auto,
                    ledger_closing,
                    match_results,
                    bank_name=st.session_state.bank_name
                )
                output_bytes = export_to_excel_bank(working_paper, recon_statement, bank_df, ledger_df, match_results)
                data_to_save = {
                    "output_bytes": output_bytes,
                    "output_filename": f"bank_reconciliation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    "match_count": len(match_results['matches']),
                    "working_paper": working_paper,
                    "recon_statement": recon_statement,
                    "match_results": match_results,
                    "bank_name": st.session_state.bank_name,
                }
                save_recon_data(data_to_save)
                st.session_state.bank_df = bank_df
                st.session_state.ledger_df = ledger_df
                st.session_state.match_results = match_results
                st.session_state.working_paper = working_paper
                st.session_state.recon_statement = recon_statement
                st.session_state.output_bytes = output_bytes
                st.session_state.output_filename = data_to_save["output_filename"]
                st.session_state.reconciliation_done = True
                st.session_state.chat_history = []
                st.success(f"✅ Bank Reconciliation complete! {len(match_results['matches'])} transactions matched.")
                st.info("📁 Results saved.")
                safe_rerun()
            except Exception as e:
                st.error(f"Error: {e}")
                import traceback
                st.code(traceback.format_exc())

    if st.session_state.reconciliation_done:
        st.markdown("---")
        st.info("💾 **Results from last bank reconciliation are shown below.**")
        match_results = st.session_state.match_results
        working_paper = st.session_state.working_paper
        recon_statement = st.session_state.recon_statement

        metric_col1, metric_col2, metric_col3, metric_col4 = st.columns(4)
        with metric_col1:
            st.metric("Matched", len(match_results['matches']))
        with metric_col2:
            st.metric("Unmatched Ledger", len(match_results['unmatched_ledger_credits']) + len(match_results['unmatched_ledger_debits']))
        with metric_col3:
            st.metric("Unmatched Bank", len(match_results['unmatched_bank_credits']) + len(match_results['unmatched_bank_debits']))
        with metric_col4:
            diff = recon_statement['difference']
            st.metric("Difference", f"{diff:,.2f}", delta="Zero" if abs(diff) < 0.01 else "Check")

        st.markdown("---")
        st.subheader("📋 Working Paper Preview")
        display_cols = ['SECTION', 'LEDGER_DATE', 'LEDGER_DESC', 'LEDGER_REF', 'LEDGER_AMOUNT', 
                        'MATCH_STATUS', 'BANK_AMOUNT', 'BANK_DATE', 'BANK_REF', 'BANK_DESC']
        wp_preview = working_paper[display_cols].head(30).copy()
        for c in ['LEDGER_DATE', 'BANK_DATE']:
            wp_preview[c] = wp_preview[c].apply(lambda x: format_date(x) if pd.notna(x) else '')
        for c in ['LEDGER_AMOUNT', 'BANK_AMOUNT']:
            wp_preview[c] = wp_preview[c].apply(lambda x: f"{float(x):,.2f}" if pd.notna(x) and x != '' else '')
        st.dataframe(wp_preview, use_container_width=True)

        st.markdown("---")
        st.subheader("📄 Clean Bank Reconciliation")
        bank_display = recon_statement.get('bank_name', '') or st.session_state.bank_name or 'Unknown Bank'
        recon_preview = [
            {"Step": f"Bank: {bank_display}", "Amount": ""},
            {"Step": "Opening Bank Balance", "Amount": f"{recon_statement['opening_balance']:,.2f}"},
            {"Step": "Closing Balance per Bank", "Amount": f"{recon_statement['bank_closing_balance']:,.2f}"},
        ]
        for _, item in recon_statement['recon_items'].iterrows():
            recon_preview.append({"Step": item['description'], "Amount": f"{item['adjustment']:,.2f}"})
        recon_preview += [
            {"Step": "TOTAL ADJUSTMENTS", "Amount": f"{recon_statement['total_adjustment']:,.2f}"},
            {"Step": "ADJUSTED BANK BALANCE", "Amount": f"{recon_statement['adjusted_balance']:,.2f}"},
            {"Step": "LEDGER CLOSING BALANCE", "Amount": f"{recon_statement['ledger_balance']:,.2f}"},
            {"Step": "UNRECONCILED DIFFERENCE", "Amount": f"{recon_statement['difference']:,.2f}"},
        ]
        st.dataframe(pd.DataFrame(recon_preview), use_container_width=True)

        st.subheader("🔎 Why the Difference Exists")
        diagnostic = pd.DataFrame([
            {"Diagnostic": "Opening Bank Balance", "Amount": recon_statement['opening_balance']},
            {"Diagnostic": "Bank Movement", "Amount": recon_statement['bank_movement']},
            {"Diagnostic": "Ledger Movement", "Amount": recon_statement['ledger_movement']},
            {"Diagnostic": "Movement Difference", "Amount": recon_statement['movement_difference']},
            {"Diagnostic": "Final Difference", "Amount": recon_statement['difference']},
        ])
        st.dataframe(diagnostic, use_container_width=True)

        if abs(recon_statement['difference']) < 0.01:
            st.success("✅ Reconciled!")
        else:
            st.warning("⚠️ Investigation required.")

        # AI Assistant for bank
        st.markdown("---")
        st.subheader("💬 Chipo – Your Centre of Intelligence")
        api_key = None
        try:
            api_key = st.secrets.get("GROQ_API_KEY", None) if hasattr(st, "secrets") else None
        except:
            pass
        if not api_key:
            st.info("ℹ️ No Groq API key found. Using rule‑based assistant.")
        else:
            st.success("✅ AI enabled using Groq (free)")

        for msg in st.session_state.chat_history:
            with st.chat_message(msg["role"]):
                st.markdown(msg["content"])

        if prompt := st.chat_input("Ask Chipo about the bank reconciliation..."):
            st.session_state.chat_history.append({"role": "user", "content": prompt})
            with st.chat_message("user"):
                st.markdown(prompt)
            with st.chat_message("assistant"):
                with st.spinner("Chipo is thinking..."):
                    context = get_ai_context(match_results, recon_statement)
                    response = get_ai_response(prompt, context, api_key)
                    st.markdown(response)
                    st.session_state.chat_history.append({"role": "assistant", "content": response})

        col_q1, col_q2, col_q3, col_q4 = st.columns(4)
        with col_q1:
            if st.button("❓ Why difference?"):
                prompt = "Why is there a difference?"
                st.session_state.chat_history.append({"role": "user", "content": prompt})
                context = get_ai_context(match_results, recon_statement)
                response = get_ai_response(prompt, context, api_key)
                st.session_state.chat_history.append({"role": "assistant", "content": response})
                safe_rerun()
        with col_q2:
            if st.button("📊 Unmatched count"):
                prompt = "How many unmatched transactions are there?"
                st.session_state.chat_history.append({"role": "user", "content": prompt})
                context = get_ai_context(match_results, recon_statement)
                response = get_ai_response(prompt, context, api_key)
                st.session_state.chat_history.append({"role": "assistant", "content": response})
                safe_rerun()
        with col_q3:
            if st.button("🔍 What to investigate first?"):
                prompt = "What should I investigate first?"
                st.session_state.chat_history.append({"role": "user", "content": prompt})
                context = get_ai_context(match_results, recon_statement)
                response = get_ai_response(prompt, context, api_key)
                st.session_state.chat_history.append({"role": "assistant", "content": response})
                safe_rerun()
        with col_q4:
            if st.button("🧹 Clear Chat"):
                st.session_state.chat_history = []
                safe_rerun()

# =========================
# Supplier Reconciliation Tab
# =========================
with tab_supplier:
    st.markdown(
        """
        <div class="hero" style="text-align:center; border: none; background: transparent; padding: 10px 0;">
            <div class="title">🔄 Supplier Reconciliation</div>
            <div class="subtitle">Upload Supplier Statement and Vendor Ledger – fully automatic, no manual mapping needed.</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    col_s1, col_s2 = st.columns(2)
    with col_s1:
        supplier_file = st.file_uploader(
            "Supplier Statement / Invoice List",
            type=["xlsx", "xls"],
            help="Upload the supplier's statement or invoice list.",
            key="supplier_file"
        )
    with col_s2:
        vendor_ledger_file = st.file_uploader(
            "Vendor Ledger",
            type=["xlsx", "xls"],
            help="Upload your vendor ledger extract.",
            key="vendor_ledger_file"
        )

    tolerance = st.slider("Amount tolerance", 0.0, 1.0, 0.01, 0.01, help="Maximum difference for a match.", key="supplier_tolerance")

    if st.button("🚀 Run Supplier Reconciliation", use_container_width=True, key="run_supplier_recon"):
        if not supplier_file or not vendor_ledger_file:
            st.error("Please upload both files.")
        else:
            with st.spinner("Processing..."):
                try:
                    result, error = reconcile_supplier(supplier_file, vendor_ledger_file, tolerance)
                    if error:
                        st.error(f"Error: {error}")
                    else:
                        st.success(f"✅ Reconciliation complete! {result['summary']['matched']} matches.")
                        st.markdown("#### Working Paper Preview")
                        st.dataframe(result['working_paper'], use_container_width=True)
                        # Download
                        excel_data = export_to_excel_supplier(result['working_paper'], result['match_results'])
                        st.download_button(
                            "📥 Download Supplier Reconciliation Report",
                            data=excel_data,
                            file_name=f"supplier_reconciliation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                        )
                except Exception as e:
                    st.error(f"Error: {e}")
                    import traceback
                    st.code(traceback.format_exc())

# =========================
# Footer with logout
# =========================
st.markdown("")
logout_c1, logout_c2, logout_c3 = st.columns([1,1,1])
with logout_c2:
    if st.button("Logout", use_container_width=True):
        logout()
